import io
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer


def _cell_value(value):
    if value is None:
        return ''
    return str(value)


def build_excel_response(title, columns, rows, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    ws.append([label for label, _ in columns])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        ws.append([_cell_value(row.get(key)) for _, key in columns])

    for i in range(1, len(columns) + 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 20

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    return response


def build_pdf_response(title, columns, rows, filename):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    elements = [Paragraph(title, styles['Title']), Spacer(1, 12)]

    data = [[label for label, _ in columns]]
    for row in rows:
        data.append([_cell_value(row.get(key)) for _, key in columns])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#333333')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f2f2f2')]),
    ]))
    elements.append(table)
    doc.build(elements)

    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
    return response
